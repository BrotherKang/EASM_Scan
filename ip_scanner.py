#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
EASM 外部攻擊面掃描系統 - 整合優化版
整合兩腳本優點：先探開放 port → 再深度漏洞掃描
支援並行、多工作表美化報告、服務風險評估、地理位置等
"""

import nmap
import re
import sys
import time
import requests
from datetime import datetime
from pathlib import Path
from concurrent.futures import ThreadPoolExecutor, as_completed
from threading import Lock
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

class EnhancedIPScanner:
    def __init__(self, ip_list_file, output_dir="scan_results", max_workers=5):
        self.ip_list_file = ip_list_file
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(exist_ok=True)
        self.timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        self.max_workers = max_workers
        self.nm = nmap.PortScanner()
        self.geo_cache = {}
        self.lock = Lock()

        self.severity_mapping = {
            'CRITICAL': {'level': 5, 'color': 'C00000'},
            'HIGH':     {'level': 4, 'color': 'FF0000'},
            'MEDIUM':   {'level': 3, 'color': 'FFC000'},
            'LOW':      {'level': 2, 'color': 'FFFF00'},
            'INFO':     {'level': 1, 'color': '00B0F0'}
        }

    def load_ip_list(self):
        try:
            with open(self.ip_list_file, 'r') as f:
                ips = [line.strip() for line in f if line.strip() and not line.startswith('#')]
            print(f"[+] 載入 {len(ips)} 個 IP")
            return ips
        except Exception as e:
            print(f"[-] 讀取失敗: {e}")
            sys.exit(1)

    def get_geo_info(self, ip):
        if ip in self.geo_cache:
            return self.geo_cache[ip]
        try:
            resp = requests.get(f"http://ip-api.com/json/{ip}?fields=status,country,city,isp,as", timeout=6).json()
            if resp.get('status') == 'success':
                info = {
                    'country': resp.get('country', 'Unknown'),
                    'city': resp.get('city', 'Unknown'),
                    'isp': resp.get('isp', 'Unknown'),
                    'asn': resp.get('as', 'Unknown')
                }
                self.geo_cache[ip] = info
                return info
        except: pass
        default = {'country': 'Unknown', 'city': 'Unknown', 'isp': 'Unknown', 'asn': 'Unknown'}
        self.geo_cache[ip] = default
        return default

    def scan_single_ip(self, ip):
        print(f"\n[*] 掃描 {ip} ...")
        result = {
            'ip': ip,
            'hostname': '',
            'status': 'down',
            'ports': [],
            'vulnerabilities': [],
            'os': 'Unknown'
        }

        geo = self.get_geo_info(ip)
        result.update(geo)

        try:
            # 第一階段：快速探測開放 port（使用預設 top 1000 + 關鍵高風險 port）
            print(f" > 階段1: 探測開放 port")
#            self.nm.scan(hosts=ip, arguments='--open -T4 --top-ports 1000 -p 22,80,443,8080,8443,3389,445,1433,3306,5432,6379')
            self.nm.scan(hosts=ip, arguments='--open -T4 -p-')

            if ip not in self.nm.all_hosts() or self.nm[ip].state() != 'up':
                result['status'] = 'down'
                result['vulnerabilities'].append({
                    'port': 'N/A', 'service': 'host', 'script': 'scan',
                    'severity': 'INFO', 'description': '主機無回應或被防火牆阻擋',
                    'cve': 'N/A', 'recommendation': '確認網路連通性與防火牆規則'
                })
                return result

            result['status'] = 'up'
            open_ports = []
            if 'tcp' in self.nm[ip]:
                open_ports = list(self.nm[ip]['tcp'].keys())

            if not open_ports:
                result['vulnerabilities'].append({
                    'port': 'N/A', 'service': 'host', 'script': 'scan',
                    'severity': 'INFO', 'description': '無開放 port',
                    'cve': 'N/A', 'recommendation': '安全狀態良好'
                })
                return result

            print(f" > 發現開放 port: {', '.join(map(str, sorted(open_ports)))}")

            # 第二階段：針對開放 port 進行深度掃描
            ports_str = ','.join(map(str, open_ports))
            web_ports = {'80', '443', '8080', '8443'}
            has_web = any(str(p) in web_ports for p in open_ports)

            scripts = ["vulners"]
            if has_web:
                scripts.extend(["ssl-enum-ciphers", "http-security-headers"])

            script_arg = ','.join(scripts)
            args = f'-sV --version-intensity 9 --script {script_arg} -p {ports_str}'

            print(f" > 階段2: 深度漏洞與服務分析")
            self.nm.scan(hosts=ip, arguments=args)

            # 解析結果
            for port in open_ports:
                if port not in self.nm[ip]['tcp']:
                    continue
                port_data = self.nm[ip]['tcp'][port]
                service = port_data.get('name', 'unknown')
                product = port_data.get('product', '')
                version = port_data.get('version', '')

                port_info = {
                    'port': str(port),
                    'protocol': 'tcp',
                    'state': port_data.get('state', 'open'),
                    'service': service,
                    'product': product,
                    'version': version,
                    'scripts': []
                }
                result['ports'].append(port_info)

                # vulners 原始輸出
                if 'script' in port_data and 'vulners' in port_data['script']:
                    vulners_out = port_data['script']['vulners'].strip()
                    if vulners_out:
                        cves = re.findall(r'CVE-\d{4}-\d{4,7}', vulners_out)
                        max_cvss = max(self.extract_cvss(vulners_out) or [0])
                        if max_cvss >= 9.0:
                            severity = 'CRITICAL'
                        elif max_cvss >= 7.0:
                            severity = 'HIGH'
                        elif max_cvss >= 4.0:
                            severity = 'MEDIUM'
                        elif max_cvss > 0:
                            severity = 'LOW'
                        else:
                            severity = 'INFO'
                        result['vulnerabilities'].append({
                            'port': str(port),
                            'service': service,
                            'script': 'vulners',
                            'severity': severity,
                            'description': vulners_out[:800] + ('...' if len(vulners_out) > 800 else ''),
                            'cve': ', '.join(sorted(set(cves))) if cves else 'N/A',
                            'recommendation': '立即更新相關軟體至最新版本，參考 Vulners 詳細資訊'
                        })

                # SSL/TLS 檢查
                if 'script' in port_data and 'ssl-enum-ciphers' in port_data['script']:
                    ssl_out = port_data['script']['ssl-enum-ciphers']
                    issues = []
                    if any(p in ssl_out for p in ["SSLv2", "SSLv3"]):
                        issues.append("極高風險：支援 SSLv2/v3")
                    if any(p in ssl_out for p in ["TLSv1.0", "TLSv1.1"]):
                        issues.append("過時協議：TLS 1.0/1.1")
                    if issues:
                        result['vulnerabilities'].append({
                            'port': str(port),
                            'service': 'https' if port in [443,8443] else 'http',
                            'script': 'ssl-enum-ciphers',
                            'severity': 'HIGH' if "極高風險" in ' '.join(issues) else 'MEDIUM',
                            'description': ' | '.join(issues),
                            'cve': 'N/A',
                            'recommendation': '停用弱協議，僅支援 TLS 1.2+，使用強加密套件'
                        })

                # HSTS 檢查
                if 'script' in port_data and 'http-security-headers' in port_data['script']:
                    headers_out = port_data['script']['http-security-headers']
                    if 'Strict-Transport-Security' not in headers_out:
                        result['vulnerabilities'].append({
                            'port': str(port),
                            'service': 'http',
                            'script': 'http-security-headers',
                            'severity': 'MEDIUM',
                            'description': '未啟用 HSTS (Strict-Transport-Security)',
                            'cve': 'N/A',
                            'recommendation': '在 Web 伺服器設定 HSTS 標頭，強制瀏覽器使用 HTTPS'
                        })

                # 服務基礎風險評估（從原腳本移植）
                service_risk = self.assess_service_risk(str(port), service, product, version)
                if service_risk:
                    result['vulnerabilities'].append(service_risk)

            # 主機名稱
            if 'hostnames' in self.nm[ip] and self.nm[ip]['hostnames']:
                result['hostname'] = self.nm[ip]['hostnames'][0].get('name', '')

            # OS 偵測
            if 'osmatch' in self.nm[ip] and self.nm[ip]['osmatch']:
                result['os'] = self.nm[ip]['osmatch'][0].get('name', 'Unknown')

        except Exception as e:
            print(f"[-] {ip} 掃描異常: {e}")
            result['vulnerabilities'].append({
                'port': 'N/A', 'service': 'error', 'script': 'exception',
                'severity': 'INFO', 'description': str(e), 'cve': 'N/A',
                'recommendation': '檢查網路或防火牆'
            })

        return result

    def extract_cvss(self, vulners_output):
        scores = []
        for line in vulners_output.splitlines():
            m = re.search(r'\b(\d+\.\d)\b', line)
            if m:
                try:
                    scores.append(float(m.group(1)))
                except:
                    pass
        return scores

    def assess_service_risk(self, port, service, product, version):
        """評估服務層級的風險（從原腳本完整移植）"""
        service_lower = service.lower()
       
        # 檢查 SSH 特殊邏輯
        if service_lower == 'ssh':
            version_str = version.lower() if version else ''
            product_str = product.lower() if product else ''
            
            # 已知 patched 的 Ubuntu 版本（backport）
            patched_ubuntu_patterns = [
                'ubuntu-3ubuntu13.3', 'ubuntu-3ubuntu13.4',  # 24.04
                'ubuntu-3ubuntu0.10',  # 22.04
                # 可再加其他發行版 patched 版本
            ]
            
            if any(patched in version_str or patched in product_str for patched in patched_ubuntu_patterns):
                # 已 patched，降為 LOW 並移除 CVE
                return {
                    'port': port,
                    'service': service,
                    'script': 'service-risk-assessment',
                    'severity': 'LOW',
                    'description': f'SSH 服務在 port {port} 開啟 {f"({product} {version})" if product else ""} (已包含 CVE-2024-6387 補丁)',
                    'cve': 'N/A',
                    'recommendation': '停用密碼登入，僅允許金鑰認證，限制允許登入的使用者和 IP'
                }
            elif '9.8' in version_str or '9.9' in version_str:  # 新版已修
                return {
                    'port': port,
                    'service': service,
                    'script': 'service-risk-assessment',
                    'severity': 'LOW',
                    'description': f'SSH 服務在 port {port} 開啟 {f"({product} {version})" if product else ""} (不受 CVE-2024-6387 影響)',
                    'cve': 'N/A',
                    'recommendation': '保持最新版本，停用密碼登入'
                }
            else:
                # 潛在 vulnerable
                return {
                    'port': port,
                    'service': service,
                    'script': 'service-risk-assessment',
                    'severity': 'CRITICAL',
                    'description': f'SSH 服務在 port {port} 開啟 ({product} {version}) - 可能易受 CVE-2024-6387 (regreSSHion) RCE 影響',
                    'cve': 'CVE-2024-6387',
                    'recommendation': '立即更新 OpenSSH 至 9.8p1 以上，或確認發行版已 backport 補丁；臨時減緩措施：設定 sshd_config LoginGraceTime 0 並重啟 sshd'
                }
       
        # 高風險服務定義
        high_risk_services = {
            'telnet': {
                'severity': 'CRITICAL',
                'description': f'Telnet 服務在 port {port} 開啟，使用明文傳輸，可被竊聽',
                'cve': 'N/A',
                'recommendation': '立即停用 Telnet，改用 SSH (port 22) 進行加密遠端連線'
            },
            'ftp': {
                'severity': 'HIGH',
                'description': f'FTP 服務在 port {port} 開啟，使用明文傳輸帳密',
                'cve': 'N/A',
                'recommendation': '停用 FTP，改用 SFTP 或 FTPS。如需繼續使用，啟用 TLS 加密 (FTPS)'
            },
            'http': {
                'severity': 'MEDIUM',
                'description': f'HTTP 服務在 port {port} 未加密，資料可被中間人攔截',
                'cve': 'N/A',
                'recommendation': '啟用 HTTPS，取得並安裝 SSL/TLS 憑證，強制重導向至 HTTPS'
            },
            'smb': {
                'severity': 'HIGH',
                'description': f'SMB 服務在 port {port} 對外開放，可能遭受 EternalBlue 等攻擊',
                'cve': 'CVE-2017-0144, CVE-2017-0145',
                'recommendation': '限制 SMB 僅內網存取，停用 SMBv1，啟用簽章驗證，更新至最新版本'
            },
            'microsoft-ds': {
                'severity': 'HIGH',
                'description': f'Microsoft-DS (SMB) 服務在 port {port} 對外開放',
                'cve': 'CVE-2017-0144, CVE-2017-0145',
                'recommendation': '限制 SMB 僅內網存取，停用 SMBv1，啟用簽章驗證'
            },
            'netbios-ssn': {
                'severity': 'MEDIUM',
                'description': f'NetBIOS 服務在 port {port} 開啟，可能洩露系統資訊',
                'cve': 'N/A',
                'recommendation': '停用 NetBIOS，或限制僅內網存取'
            },
            'rdp': {
                'severity': 'HIGH',
                'description': f'RDP 服務在 port {port} 對外開放，常遭暴力破解攻擊',
                'cve': 'CVE-2019-0708 (BlueKeep)',
                'recommendation': '啟用網路層級驗證 (NLA)，使用多因素驗證 (MFA)，限制允許連線的 IP，更新至最新版本'
            },
            'ms-wbt-server': {
                'severity': 'HIGH',
                'description': f'RDP 服務在 port {port} 對外開放',
                'cve': 'CVE-2019-0708 (BlueKeep)',
                'recommendation': '啟用 NLA，使用 MFA，限制 IP 白名單'
            },
            'mysql': {
                'severity': 'MEDIUM',
                'description': f'MySQL 資料庫在 port {port} 對外曝露',
                'cve': 'N/A',
                'recommendation': '限制 MySQL 僅 localhost 或內網存取，使用強密碼，定期更新版本'
            },
            'ms-sql-s': {
                'severity': 'MEDIUM',
                'description': f'MS SQL Server 在 port {port} 對外曝露',
                'cve': 'N/A',
                'recommendation': '限制僅內網存取，啟用 Windows 驗證，加密連線，定期更新'
            },
            'postgresql': {
                'severity': 'MEDIUM',
                'description': f'PostgreSQL 資料庫在 port {port} 對外曝露',
                'cve': 'N/A',
                'recommendation': '限制僅內網或特定 IP 存取，使用強密碼，啟用 SSL 連線'
            },
            'mongodb': {
                'severity': 'HIGH',
                'description': f'MongoDB 在 port {port} 對外曝露，可能未啟用驗證',
                'cve': 'N/A',
                'recommendation': '啟用驗證機制，限制僅內網存取，使用防火牆規則'
            },
            'redis': {
                'severity': 'HIGH',
                'description': f'Redis 在 port {port} 對外曝露，預設無密碼保護',
                'cve': 'N/A',
                'recommendation': '設定強密碼 (requirepass)，綁定至 localhost，停用危險指令'
            },
            'vnc': {
                'severity': 'HIGH',
                'description': f'VNC 服務在 port {port} 開啟，可能使用弱加密',
                'cve': 'N/A',
                'recommendation': '使用 SSH 隧道加密 VNC 流量，或改用 RDP/其他加密遠端方案'
            },
            'smtp': {
                'severity': 'MEDIUM',
                'description': f'SMTP 服務在 port {port} 開啟，需檢查是否為開放轉發',
                'cve': 'N/A',
                'recommendation': '停用開放轉發 (Open Relay)，啟用 STARTTLS，設定 SPF/DKIM/DMARC'
            },
            'pop3': {
                'severity': 'MEDIUM',
                'description': f'POP3 服務在 port {port} 開啟，使用明文傳輸',
                'cve': 'N/A',
                'recommendation': '改用 POP3S (SSL/TLS 加密) 或 IMAP，停用明文 POP3'
            },
            'imap': {
                'severity': 'MEDIUM',
                'description': f'IMAP 服務在 port {port} 開啟，使用明文傳輸',
                'cve': 'N/A',
                'recommendation': '改用 IMAPS (SSL/TLS 加密)，停用明文 IMAP'
            },
            'elasticsearch': {
                'severity': 'HIGH',
                'description': f'Elasticsearch 在 port {port} 對外曝露',
                'cve': 'N/A',
                'recommendation': '限制僅內網存取，啟用 X-Pack 安全功能，使用驗證與加密'
            },
            'docker': {
                'severity': 'CRITICAL',
                'description': f'Docker API 在 port {port} 未加密對外開放，可被遠端控制',
                'cve': 'N/A',
                'recommendation': '啟用 TLS 驗證，限制僅內網存取，或使用 SSH 隧道'
            },
            'kubernetes': {
                'severity': 'CRITICAL',
                'description': f'Kubernetes API 在 port {port} 對外曝露',
                'cve': 'N/A',
                'recommendation': '啟用 RBAC，使用網路策略限制存取，啟用 TLS 驗證'
            }
        }
       
        # 檢查是否為高風險服務
        if service_lower in high_risk_services:
            risk_info = high_risk_services[service_lower]
            return {
                'port': port,
                'service': service,
                'script': 'service-risk-assessment',
                'severity': risk_info['severity'],
                'description': risk_info['description'],
                'cve': risk_info['cve'],
                'recommendation': risk_info['recommendation']
            }
       
        # 檢查常見的高風險 port
        risky_ports = {
            '21': ('FTP', 'HIGH'),
            '23': ('Telnet', 'CRITICAL'),
            '69': ('TFTP', 'HIGH'),
            '135': ('MS-RPC', 'MEDIUM'),
            '139': ('NetBIOS', 'MEDIUM'),
            '445': ('SMB', 'HIGH'),
            '1433': ('MS-SQL', 'MEDIUM'),
            '3306': ('MySQL', 'MEDIUM'),
            '3389': ('RDP', 'HIGH'),
            '5432': ('PostgreSQL', 'MEDIUM'),
            '5900': ('VNC', 'HIGH'),
            '6379': ('Redis', 'HIGH'),
            '8080': ('HTTP-Proxy', 'LOW'),
            '9200': ('Elasticsearch', 'HIGH'),
            '27017': ('MongoDB', 'HIGH')
        }
       
        if port in risky_ports and service_lower not in high_risk_services:
            service_name, severity = risky_ports[port]
            return {
                'port': port,
                'service': service or service_name,
                'script': 'port-risk-assessment',
                'severity': severity,
                'description': f'Port {port} ({service_name}) 開啟，可能存在安全風險',
                'cve': 'N/A',
                'recommendation': f'檢視 {service_name} 服務的必要性，如非必要請關閉，或限制存取來源'
            }
       
        return None

    def scan_parallel(self, ip_list):
        results = []
        with ThreadPoolExecutor(max_workers=self.max_workers) as executor:
            futures = {executor.submit(self.scan_single_ip, ip): ip for ip in ip_list}
            for future in as_completed(futures):
                try:
                    res = future.result()
                    with self.lock:
                        results.append(res)
                except Exception as e:
                    print(f"[-] 處理異常: {e}")
        return results

    def generate_report(self, scan_results):
        """產生Excel報告（從原腳本完整移植，包括所有美化）"""
        wb = openpyxl.Workbook()
       
        # 建立工作表
        ws_summary = wb.active
        ws_summary.title = "掃描摘要"
        ws_detail = wb.create_sheet("詳細結果")
        ws_vuln = wb.create_sheet("漏洞清單")
        ws_port = wb.create_sheet("Port開啟統計")
        ws_risk = wb.create_sheet("風險評估")
       
        # 樣式定義
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
       
        # === 掃描摘要工作表 ===
        summary_headers = ["項目", "數值", "說明"]
        ws_summary.append(summary_headers)
       
        total_ips = len(scan_results)
        total_vulnerabilities = sum(len(r['vulnerabilities']) for r in scan_results)
        critical_count = sum(1 for r in scan_results for v in r['vulnerabilities'] if v['severity'] == 'CRITICAL')
        high_count = sum(1 for r in scan_results for v in r['vulnerabilities'] if v['severity'] == 'HIGH')
        medium_count = sum(1 for r in scan_results for v in r['vulnerabilities'] if v['severity'] == 'MEDIUM')
        low_count = sum(1 for r in scan_results for v in r['vulnerabilities'] if v['severity'] == 'LOW')
       
        summary_data = [
            ["掃描時間", datetime.now().strftime("%Y-%m-%d %H:%M:%S"), "報告產生時間"],
            ["掃描IP總數", total_ips, "本次掃描的IP數量"],
            ["線上主機數", sum(1 for r in scan_results if r['status'] == 'up'), "狀態為up的主機"],
            ["發現漏洞總數", total_vulnerabilities, "所有漏洞總計"],
            ["🔴 危急(Critical)", critical_count, "需立即處理"],
            ["🟠 高危(High)", high_count, "優先修補"],
            ["🟡 中危(Medium)", medium_count, "排程修補"],
            ["🔵 低危(Low)", low_count, "例行維護"],
            ["開啟Port總數", sum(len(r['ports']) for r in scan_results), "所有開啟的port"],
            ["平均每主機Port數", round(sum(len(r['ports']) for r in scan_results) / max(total_ips, 1), 2), ""],
            ["平均每主機漏洞數", round(total_vulnerabilities / max(total_ips, 1), 2), ""]
        ]
       
        for row in summary_data:
            ws_summary.append(row)
       
        self.format_sheet(ws_summary, header_fill, header_font, border)
       
        # === 詳細結果工作表 ===
        detail_headers = ["IP位址", "主機名稱", "國家/地區", "城市", "電信商(ISP)", "ASN", "狀態", "作業系統", "開啟Port數", "漏洞數量", "最高風險等級", "總結建議"]
        ws_detail.append(detail_headers)
       
        for result in scan_results:
            # 計算最高風險等級
            max_severity = 'INFO'
            if result['vulnerabilities']:
                severity_levels = [self.severity_mapping.get(v['severity'], {'level': 1})['level'] for v in result['vulnerabilities']]
                max_level = max(severity_levels)
                max_severity = [k for k, v in self.severity_mapping.items() if v['level'] == max_level][0]
           
            # 彙總修復建議（去除重複，並以換行分隔）
            recommendations = set(v['recommendation'] for v in result['vulnerabilities'] if 'recommendation' in v)
            summary_recommendation = '\n'.join(recommendations) if recommendations else '無特定建議'
           
            row_data = [
                result['ip'],
                result.get('hostname', 'N/A'),
                result.get('country', 'Unknown'),
                result.get('city', 'Unknown'),
                result.get('isp', 'Unknown'),
                result.get('asn', 'Unknown'),
                result['status'],
                result['os'],
                len(result['ports']),
                len(result['vulnerabilities']),
                max_severity,
                summary_recommendation
            ]
            ws_detail.append(row_data)
           
            # 根據風險等級標色
            if max_severity in self.severity_mapping:
                color = self.severity_mapping[max_severity]['color']
                ws_detail.cell(row=ws_detail.max_row, column=11).fill = PatternFill(
                    start_color=color, end_color=color, fill_type="solid"
                )
       
        self.format_sheet(ws_detail, header_fill, header_font, border)
        ws_detail.column_dimensions['L'].width = 60  # 調整總結建議欄寬
       
        # === 漏洞清單工作表 ===
        vuln_headers = ["IP位址", "主機名稱", "Port", "服務", "漏洞嚴重程度", "檢測腳本", "CVE編號", "漏洞描述", "修復建議"]
        ws_vuln.append(vuln_headers)
       
        for result in scan_results:
            for vuln in result['vulnerabilities']:
                row_data = [
                    result['ip'],
                    result.get('hostname', 'N/A'),
                    vuln['port'],
                    vuln['service'],
                    vuln['severity'],
                    vuln['script'],
                    vuln.get('cve', 'N/A'),
                    vuln['description'],
                    vuln['recommendation']
                ]
                ws_vuln.append(row_data)
               
                # 根據嚴重程度設定顏色
                severity_color = self.severity_mapping.get(vuln['severity'], {}).get('color', 'FFFFFF')
                ws_vuln.cell(row=ws_vuln.max_row, column=5).fill = PatternFill(
                    start_color=severity_color, end_color=severity_color, fill_type="solid"
                )
       
        self.format_sheet(ws_vuln, header_fill, header_font, border)
        ws_vuln.column_dimensions['H'].width = 50
        ws_vuln.column_dimensions['I'].width = 40
       
        # === Port統計工作表 ===
        port_stats = {}
        for result in scan_results:
            for port_info in result['ports']:
                port_key = f"{port_info['port']}/{port_info['protocol']}"
                service_name = port_info['service'] or 'unknown'
               
                if port_key not in port_stats:
                    port_stats[port_key] = {'count': 0, 'service': service_name, 'ips': []}
               
                port_stats[port_key]['count'] += 1
                port_stats[port_key]['ips'].append(result['ip'])
       
        port_headers = ["Port/協議", "服務名稱", "出現次數", "曝險比例", "受影響IP"]
        ws_port.append(port_headers)
       
        for port, stats in sorted(port_stats.items(), key=lambda x: x[1]['count'], reverse=True):
            exposure_rate = f"{(stats['count'] / total_ips * 100):.1f}%" if total_ips > 0 else "0%"
            row_data = [
                port,
                stats['service'],
                stats['count'],
                exposure_rate,
                ', '.join(stats['ips'][:10]) + ('...' if len(stats['ips']) > 10 else '')
            ]
            ws_port.append(row_data)
       
        self.format_sheet(ws_port, header_fill, header_font, border)
        ws_port.column_dimensions['E'].width = 50
       
        # === 風險評估工作表 ===
        risk_headers = ["風險類型", "風險等級", "影響主機數", "建議措施"]
        ws_risk.append(risk_headers)
       
        # 高風險服務識別（從原腳本移植）
        risky_services = {
            'telnet': ('CRITICAL', '使用未加密協議'),
            'ftp': ('HIGH', '使用未加密協議'),
            'http': ('MEDIUM', '未使用加密傳輸'),
            'smb': ('HIGH', '可能遭受勒索軟體攻擊'),
            'rdp': ('HIGH', '常見暴力破解目標'),
            'mysql': ('MEDIUM', '資料庫對外曝露'),
            'mssql': ('MEDIUM', '資料庫對外曝露'),
            'mongodb': ('HIGH', '資料庫對外曝露'),
            'redis': ('HIGH', '快取資料庫對外曝露')
        }
       
        for service, (severity, desc) in risky_services.items():
            affected_hosts = [r['ip'] for r in scan_results for p in r['ports'] if service in p['service'].lower()]
            if affected_hosts:
                row_data = [
                    f"{service.upper()} 服務曝露",
                    severity,
                    len(affected_hosts),
                    f"{desc}，建議立即檢視：{', '.join(affected_hosts[:5])}"
                ]
                ws_risk.append(row_data)
               
                color = self.severity_mapping.get(severity, {}).get('color', 'FFFFFF')
                ws_risk.cell(row=ws_risk.max_row, column=2).fill = PatternFill(
                    start_color=color, end_color=color, fill_type="solid"
                )
       
        self.format_sheet(ws_risk, header_fill, header_font, border)
        ws_risk.column_dimensions['D'].width = 60
       
        # 儲存檔案
        output_file = self.output_dir / f"EASM_Enhanced_Report_{self.timestamp}.xlsx"
        wb.save(output_file)
        print(f"[+] 報告已產生: {output_file}")
       
        return output_file

    def format_sheet(self, ws, header_fill, header_font, border):
        """格式化工作表（從原腳本移植）"""
        # 標題列格式
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
       
        # 所有儲存格加上框線
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
            for cell in row:
                cell.border = border
                if cell.row > 1:
                    cell.alignment = Alignment(vertical='top', wrap_text=True)
       
        # 自動調整欄寬
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
           
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
           
            adjusted_width = min(max_length + 2, 60)
            ws.column_dimensions[column_letter].width = adjusted_width

    def run(self):
        print("=" * 70)
        print("EASM 外部攻擊面掃描系統 - 整合優化版（動態雙階段掃描）")
        print("=" * 70)

        start_time = time.time()
        ips = self.load_ip_list()
        scan_results = self.scan_parallel(ips)
       
        if not scan_results:
            print("[-] 沒有成功的掃描結果")
            return
       
        # 產生報告
        print("\n[*] 產生Excel報告...")
        report_file = self.generate_report(scan_results)
       
        elapsed_time = time.time() - start_time
       
        print("\n" + "="*60)
        print(f"掃描完成! 共掃描 {len(scan_results)} 個IP")
        print(f"總耗時: {elapsed_time:.2f} 秒")
        print(f"平均每個IP: {elapsed_time/len(scan_results):.2f} 秒")
        print(f"報告檔案: {report_file}")
        print("="*60)

if __name__ == "__main__":
    import argparse
    parser = argparse.ArgumentParser(description="整合版 EASM 掃描器")
    parser.add_argument("ip_list", help="IP 清單檔案")
    parser.add_argument("--workers", type=int, default=6, help="並行數 (預設 6)")
    parser.add_argument("--output", default="scan_results", help="輸出目錄")
    args = parser.parse_args()

    scanner = EnhancedIPScanner(args.ip_list, args.output, args.workers)
    scanner.run()
